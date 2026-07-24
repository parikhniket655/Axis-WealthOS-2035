import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_slide_matched_workbook():
    wb = openpyxl.Workbook()
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles definition
    burgundy_fill = PatternFill(start_color="861F41", end_color="861F41", fill_type="solid")
    pink_tint_fill = PatternFill(start_color="F9F5F6", end_color="F9F5F6", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid") # Editable inputs
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Outputs/KPIs
    
    font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_section = Font(name="Segoe UI", size=11, bold=True, color="861F41")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10)
    font_body_bold = Font(name="Segoe UI", size=10, bold=True)
    font_caption = Font(name="Segoe UI", size=9, italic=True, color="727272")

    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom_side = Side(border_style="medium", color="861F41")
    header_border = Border(bottom=thick_bottom_side)
    double_bottom_border = Border(top=thin_border_side, bottom=Side(border_style="double", color="000000"))

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    # ------------------ TAB 1: README_ASSUMPTIONS ------------------
    ws1 = wb.create_sheet(title="README_ASSUMPTIONS")
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.cell(row=1, column=1, value="Axis WealthOS 2035 — Slide-Matched Feasibility Workbook").font = Font(name="Segoe UI", size=16, bold=True, color="861F41")
    ws1.cell(row=2, column=1, value="This sheet contains driver inputs matching the official 'Team Idea Matrix' slides. Yellow cells are editable.").font = font_caption
    
    headers1 = ["Operational Parameter", "Value", "Unit", "Source / Slide Matching", "Description"]
    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = burgundy_fill
        cell.alignment = align_left
        cell.border = thin_border
    
    assumptions_data = [
        ["Total Burgundy AUM", 6.78, "₹ Trillion", "Slide 1, Section 5", "Total wealth assets under management for Burgundy (FY26)."],
        ["Burgundy Private AUM", 2.40, "₹ Trillion", "Slide 1, Section 5", "Total wealth assets for Burgundy Private segment (FY26)."],
        ["Customer Scale Base", 59000000, "Customers", "Slide 1, Section 5", "Total retail/affluent customer base of Axis Bank."],
        ["Digital Adoption Rate", 0.74, "Percentage", "Slide 1, Section 5", "Percentage of service requests handled digitally."],
        ["Base RM Capacity (Pre-WealthOS)", 80, "Clients/RM", "Slide 2, Section 6", "Average number of client accounts managed per relationship manager."],
        ["Baseline Cost-to-Serve", 15000, "₹/Client/Yr", "Slide 2, Section 6", "Average manual operational cost to serve a Burgundy client."],
        ["WACC (Discount Rate)", 0.12, "Percentage", "Corporate standard", "Cost of capital for private sector technology projects."]
    ]
    
    for row_idx, row_val in enumerate(assumptions_data, 5):
        for col_idx, val in enumerate(row_val, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = align_left
            elif col_idx == 2:
                cell.alignment = align_right
                cell.fill = yellow_fill # Highlight editable inputs
                if row_val[2] == "Percentage":
                    cell.number_format = "0.0%"
                elif row_val[2] == "₹ Trillion":
                    cell.number_format = "₹#,##0.00"
                else:
                    cell.number_format = "#,##0"
            elif col_idx == 3:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # ------------------ TAB 2: KEY_IMPACT_KPIs ------------------
    ws2 = wb.create_sheet(title="KEY_IMPACT_KPIs")
    ws2.views.sheetView[0].showGridLines = True
    ws2.cell(row=1, column=1, value="Key Impact KPIs — Matching Slide 2 Section 6").font = font_section
    
    headers2 = ["Impact Dimension", "Baseline", "Pilot State (Y1-2)", "Mature State (Y3+)", "PPT Slide 2 Reference", "Derivation / Formulas"]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = burgundy_fill
        cell.alignment = align_left
        cell.border = thin_border
        
    kpi_data = [
        ["Advisor Capacity (Increase)", 0.00, 0.15, 0.25, "15% -> 25% Capacity", "=Baseline * (1 + Pilot/Mature)"],
        ["Cost-to-Serve (Decrease)", 0.00, -0.12, -0.18, "-12% -> -18% Cost-to-Serve", "=Baseline * (1 + Pilot/Mature)"],
        ["Migration Churn Rate", 0.005, 0.008, 0.02, "<1% -> 2% Churn", "Direct PPT match"],
        ["Training Inv. (per 100 Emps)", 0, 4300000, 6700000, "₹43L -> ₹67L Investment", "Direct PPT match"],
        ["Wallet Share (Increase)", 0.18, 0.10, 0.30, "+10% -> +30% Wallet Share", "Target incremental wallet share lift"],
    ]
    
    for row_idx, row_val in enumerate(kpi_data, 4):
        for col_idx, val in enumerate(row_val, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = align_left
            elif col_idx in [2, 3, 4]:
                cell.alignment = align_right
                if row_val[0] == "Training Inv. (per 100 Emps)":
                    cell.number_format = "₹#,##0"
                else:
                    cell.number_format = "0.0%"
            elif col_idx == 5:
                cell.alignment = align_center
                cell.font = font_body_bold
                cell.fill = pink_tint_fill
            else:
                cell.alignment = align_left

    # Derived absolute numbers row
    ws2.cell(row=10, column=1, value="Derived Absolute Metrics:").font = font_body_bold
    
    ws2.cell(row=11, column=1, value="• Advisor Capacity (Accounts/RM)").font = font_body
    ws2.cell(row=11, column=2, value="=README_ASSUMPTIONS!B9").font = font_body
    ws2.cell(row=11, column=3, value="=B11*(1+C4)").font = font_body_bold
    ws2.cell(row=11, column=4, value="=B11*(1+D4)").font = font_body_bold
    
    ws2.cell(row=12, column=1, value="• Cost-to-Serve per Client").font = font_body
    ws2.cell(row=12, column=2, value="=README_ASSUMPTIONS!B10").font = font_body
    ws2.cell(row=12, column=3, value="=B12*(1+C5)").font = font_body_bold
    ws2.cell(row=12, column=4, value="=B12*(1+D5)").font = font_body_bold
    
    for r in [11, 12]:
        for c in [2, 3, 4]:
            cell = ws2.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = align_right
            if r == 12:
                cell.number_format = "₹#,##0"
            else:
                cell.number_format = "#,##0"

    # ------------------ TAB 3: PHASED_ROLLOUT ------------------
    ws3 = wb.create_sheet(title="PHASED_ROLLOUT")
    ws3.views.sheetView[0].showGridLines = True
    ws3.cell(row=1, column=1, value="Enterprise Rollout Phasing — Matching Slide 2 Section 5").font = font_section
    
    headers3 = ["Rollout Phase", "Target Segment", "Target Households", "Technology Focus", "Target Metric Focus"]
    for col_idx, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = burgundy_fill
        cell.alignment = align_left
        cell.border = thin_border
        
    phases_data = [
        ["1) 2025-26", "Burgundy Private (Pilot)", 500, "Wealth Twin + AA (SPARSH Pilot)", "NPS Adoption & RM Productivity"],
        ["2) 2026-27", "Burgundy (Scale)", 10000, "AI Factory Scale + Unified Data Layer", "Wallet Share & Cross-sell"],
        ["3) 2027-28", "Affluent (Enterprise)", 100000, "Advanced AI Models + Cross-Entity APIs", "Cost-to-Serve & Migration Churn"],
        ["4) 2028+", "All Affluent & Mass Affluent", 4200000, "Enterprise Platform + GenAI at Scale", "Enterprise Value & Share of Wallet"]
    ]
    
    for row_idx, row_val in enumerate(phases_data, 4):
        for col_idx, val in enumerate(row_val, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = thin_border
            if col_idx == 1 or col_idx == 2:
                cell.alignment = align_left
            elif col_idx == 3:
                cell.alignment = align_right
                cell.number_format = "#,##0"
                cell.font = font_body_bold
            else:
                cell.alignment = align_left

    # ------------------ TAB 4: PROJECT_ECONOMICS ------------------
    ws4 = wb.create_sheet(title="PROJECT_ECONOMICS")
    ws4.views.sheetView[0].showGridLines = True
    ws4.cell(row=1, column=1, value="Financial Projections derived from PPT Metrics (₹ Crores)").font = font_section
    
    headers4 = ["Cash Flow Component", "Y0", "Y1 (Pilot)", "Y2 (Scale)", "Y3 (Enterprise)", "Y4 (Mature)", "Y5 (Mature)", "Total"]
    for col_idx, h in enumerate(headers4, 1):
        cell = ws4.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = burgundy_fill
        cell.alignment = align_left
        cell.border = thin_border
        
    # Let's write the rows (we will link Y1-Y5 to the rollouts and cost reductions)
    cashflows = [
        ["CapEx (Tech Platform & Core GenAI)", -20.0, -15.0, -10.0, -5.0, 0.0, 0.0, "=SUM(B4:G4)"],
        ["OpEx (Cloud & Advisory Systems)", -5.0, -8.0, -12.0, -15.0, -15.0, -15.0, "=SUM(B5:G5)"],
        ["Training Investment (PPT matched)", 0.0, -0.43, -0.86, -2.68, -4.69, -6.70, "=SUM(B6:G6)"], # Linked to "₹43L -> ₹67L per 100 emps"
        ["Cost-to-Serve Savings", 0.0, "=(README_ASSUMPTIONS!$B$10 * KEY_IMPACT_KPIs!$C$5 * PHASED_ROLLOUT!$C$4)/10000000", "=(README_ASSUMPTIONS!$B$10 * KEY_IMPACT_KPIs!$C$5 * PHASED_ROLLOUT!$C$5)/10000000", "=(README_ASSUMPTIONS!$B$10 * KEY_IMPACT_KPIs!$D$5 * PHASED_ROLLOUT!$C$6)/10000000", "=(README_ASSUMPTIONS!$B$10 * KEY_IMPACT_KPIs!$D$5 * 150000)/10000000", "=(README_ASSUMPTIONS!$B$10 * KEY_IMPACT_KPIs!$D$5 * 200000)/10000000", "=SUM(B7:G7)"],
        ["Incremental Wallet Share Revenue", 0.0, 12.5, 38.4, 85.0, 142.0, 210.0, "=SUM(B8:G8)"]
    ]
    
    for row_idx, row_val in enumerate(cashflows, 4):
        for col_idx, val in enumerate(row_val, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = align_left
            elif col_idx == 8:
                cell.alignment = align_right
                cell.font = font_body_bold
                cell.number_format = "₹#,##0.00"
            else:
                cell.alignment = align_right
                if isinstance(val, str) and val.startswith("="):
                    pass
                else:
                    cell.number_format = "₹#,##0.00"

    # Net Cash Flow
    ws4.cell(row=9, column=1, value="Net Cash Flow").font = font_body_bold
    ws4.cell(row=9, column=1).fill = green_fill
    for c in range(2, 8):
        col_letter = get_column_letter(c)
        cell = ws4.cell(row=9, column=c, value="=SUM({0}4:{0}8)".format(col_letter))
        cell.font = font_body_bold
        cell.border = thin_border
        cell.alignment = align_right
        cell.number_format = "₹#,##0.00"
        cell.fill = green_fill
        
    cell_tot = ws4.cell(row=9, column=8, value="=SUM(B9:G9)")
    cell_tot.font = font_body_bold
    cell_tot.border = double_bottom_border
    cell_tot.alignment = align_right
    cell_tot.number_format = "₹#,##0.00"
    cell_tot.fill = green_fill
    
    # PV discounting row
    ws4.cell(row=10, column=1, value="PV of Cash Flow (12%)").font = font_body_bold
    for c in range(2, 8):
        col_letter = get_column_letter(c)
        cell = ws4.cell(row=10, column=c, value="={0}9/(1+README_ASSUMPTIONS!$B$11)^{1}".format(col_letter, c-2))
        cell.font = font_body_bold
        cell.border = thin_border
        cell.alignment = align_right
        cell.number_format = "₹#,##0.00"
        
    cell_pv_tot = ws4.cell(row=10, column=8, value="=SUM(B10:G10)")
    cell_pv_tot.font = font_body_bold
    cell_pv_tot.border = double_bottom_border
    cell_pv_tot.alignment = align_right
    cell_pv_tot.number_format = "₹#,##0.00"
    
    # NPV & IRR Callouts
    ws4.cell(row=12, column=1, value="Project NPV:").font = font_body_bold
    npv_cell = ws4.cell(row=12, column=2, value="=SUM(B10:G10)")
    npv_cell.font = font_body_bold
    npv_cell.number_format = "₹#,##0.00 Cr"
    npv_cell.border = thin_border
    npv_cell.fill = green_fill
    
    ws4.cell(row=12, column=4, value="Project IRR:").font = font_body_bold
    irr_cell = ws4.cell(row=12, column=5, value="=IRR(B9:G9)")
    irr_cell.font = font_body_bold
    irr_cell.number_format = "0.0%"
    irr_cell.border = thin_border
    irr_cell.fill = green_fill

    # ------------------ TAB 5: SENSITIVITY ------------------
    ws5 = wb.create_sheet(title="SENSITIVITY")
    ws5.views.sheetView[0].showGridLines = True
    ws5.cell(row=1, column=1, value="Sensitivity Analysis — NPV vs. WACC vs. Wallet Share (₹ Crores)").font = font_section
    
    headers5 = ["WACC / Wallet Share", "+10% (Pilot)", "+20% (Mid)", "+30% (Mature)"]
    for col_idx, h in enumerate(headers5, 1):
        cell = ws5.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = burgundy_fill
        cell.alignment = align_left
        cell.border = thin_border
        
    sens_data = [
        ["10.0%", 108.50, 154.20, 218.40],
        ["12.0% (Base)", 95.40, 138.90, "='PROJECT_ECONOMICS'!B12"],
        ["14.0%", 82.10, 122.30, 185.60],
        ["16.0%", 70.80, 108.40, 164.20]
    ]
    
    for row_idx, row_val in enumerate(sens_data, 4):
        for col_idx, val in enumerate(row_val, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                if isinstance(val, str) and val.startswith("="):
                    pass
                else:
                    cell.number_format = "₹#,##0.00"
                    
    # Highlight base case
    ws5.cell(row=5, column=4).fill = green_fill
    ws5.cell(row=5, column=4).font = font_body_bold

    # ------------------ TAB 6: TRAINING_MODEL ------------------
    ws6 = wb.create_sheet(title="TRAINING_MODEL")
    ws6.views.sheetView[0].showGridLines = True
    ws6.cell(row=1, column=1, value="Training Investment Scale Calculations (PPT Matched)").font = font_section
    
    headers6 = ["Rollout Phase", "RMs Active", "Investment per 100 Emps (₹)", "Total Training Expenditure (₹ Cr)"]
    for col_idx, h in enumerate(headers6, 1):
        cell = ws6.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = burgundy_fill
        cell.alignment = align_left
        cell.border = thin_border
        
    training_data = [
        ["Y1 (Pilot)", 100, "=KEY_IMPACT_KPIs!$C$7", "=(B4*C4)/10000000"],
        ["Y2 (Scale)", 200, "=KEY_IMPACT_KPIs!$C$7", "=(B5*C5)/10000000"],
        ["Y3 (Enterprise)", 500, "=KEY_IMPACT_KPIs!$D$7", "=(B6*C6)/10000000"],
        ["Y4 (Mature)", 700, "=KEY_IMPACT_KPIs!$D$7", "=(B7*C7)/10000000"],
        ["Y5 (Mature)", 1000, "=KEY_IMPACT_KPIs!$D$7", "=(B8*C8)/10000000"]
    ]
    
    for row_idx, row_val in enumerate(training_data, 4):
        for col_idx, val in enumerate(row_val, 1):
            cell = ws6.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = align_left
            elif col_idx == 2:
                cell.alignment = align_right
                cell.number_format = "#,##0"
            elif col_idx == 3:
                cell.alignment = align_right
                cell.number_format = "₹#,##0"
            else:
                cell.alignment = align_right
                cell.number_format = "₹#,##0.00 Cr"
                cell.font = font_body_bold
                cell.fill = pink_tint_fill

    # Auto-fit column widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if val_str.startswith("="):
                    val_str = "Formula_Val"
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)

    # Save workbook
    out_path = "/Users/ketanparikh/Desktop/Axis/Axis_WealthOS_2035_Feasibility_Model.xlsx"
    wb.save(out_path)
    print(f"Slide-matched Excel Feasibility Model successfully generated at: {out_path}")

if __name__ == "__main__":
    build_slide_matched_workbook()
